from __future__ import annotations

import io
import os
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from typing import List, Optional, Any, Tuple

from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

from starlette.middleware.sessions import SessionMiddleware
from starlette.staticfiles import StaticFiles
import secrets


# =========================
# Config
# =========================
APP_TITLE = "Cotizador - Seguritec Tarija"
EMPRESA_NOMBRE = "Seguritec Tarija"
EMPRESA_TELF = "70218010"
IVA_RATE = 0.13  # 13%

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
DB_PATH = os.path.join(BASE_DIR, "app.db")
EXCEL_PATH = os.path.join(DATA_DIR, "precios.xlsx")

# Render / Postgres
DATABASE_URL = os.environ.get("DATABASE_URL")  # en Render ponla como DATABASE_URL

ADMIN_USER = os.environ.get("ADMIN_USER", "seguritec")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "cambia_esto")
SESSION_SECRET = os.environ.get("SESSION_SECRET", secrets.token_urlsafe(32))

# Logo (opcional). Si existe: static/logo.png
LOGO_PATH = os.environ.get("LOGO_PATH", os.path.join(STATIC_DIR, "logo.png"))

app = FastAPI(title=APP_TITLE)
templates = Jinja2Templates(directory=TEMPLATES_DIR)

app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


# =========================
# DB backend selector
# =========================
IS_POSTGRES = bool(DATABASE_URL)

def _pg_connect():
    import psycopg
    from psycopg.rows import dict_row

    url = DATABASE_URL
    if url and "sslmode=" not in url:
        sep = "&" if "?" in url else "?"
        url = url + f"{sep}sslmode=require"

    return psycopg.connect(url, row_factory=dict_row)

def db_connect():
    """
    Devuelve conexión (sqlite o postgres).
    """
    if IS_POSTGRES:
        con = _pg_connect()
        return con
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def db_exec(sql: str, params: Tuple[Any, ...] = ()):
    """
    Ejecuta SQL de forma portable.
    En Postgres: placeholders %s
    En SQLite: placeholders ?
    """
    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute(sql, params)
        con.commit()
        return cur
    finally:
        try:
            con.close()
        except Exception:
            pass

def db_fetchone(sql: str, params: Tuple[Any, ...] = ()):
    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute(sql, params)
        return cur.fetchone()
    finally:
        try:
            con.close()
        except Exception:
            pass

def db_fetchall(sql: str, params: Tuple[Any, ...] = ()):
    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute(sql, params)
        return cur.fetchall()
    finally:
        try:
            con.close()
        except Exception:
            pass

def psql(sql: str) -> str:
    """
    Convierte placeholders ? (sqlite style) a %s (postgres style).
    Es simple y suficiente para este proyecto (no uses ? en strings).
    """
    if not IS_POSTGRES:
        return sql
    return sql.replace("?", "%s")


# =========================
# Auth helpers
# =========================
def is_logged_in(request: Request) -> bool:
    return bool(request.session.get("auth"))

def require_login(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login", status_code=303)
    return None


# =========================
# Catalog types
# =========================
@dataclass
class Product:
    sku: str
    categoria: str
    nombre: str
    unidad: str
    precio_bs: float
    activo: int = 1


# =========================
# Excel helpers (solo para import)
# =========================
def _norm_header(x) -> str:
    return str(x).strip().lower() if x is not None else ""

def _to_float(x, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        s = str(x).strip().replace(",", ".")
        try:
            return float(s)
        except Exception:
            return default

def _to_int(x, default: int = 0) -> int:
    try:
        if x is None:
            return default
        return int(float(x))
    except Exception:
        return default

def read_products_from_excel() -> List[Product]:
    if not os.path.exists(EXCEL_PATH):
        return []

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "productos" not in wb.sheetnames:
        raise ValueError("El Excel debe tener una hoja llamada 'productos'.")

    ws = wb["productos"]

    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        h = _norm_header(cell.value)
        if h:
            headers[h] = col_idx

    required = {"sku", "categoria", "nombre", "unidad", "precio_bs", "activo"}
    missing = required - set(headers.keys())
    if missing:
        raise ValueError(f"Faltan columnas en Excel: {', '.join(sorted(missing))}")

    products: List[Product] = []
    for r in range(2, ws.max_row + 1):
        activo = _to_int(ws.cell(r, headers["activo"]).value, default=0)
        if activo != 1:
            continue

        sku = str(ws.cell(r, headers["sku"]).value or "").strip()
        categoria = str(ws.cell(r, headers["categoria"]).value or "").strip()
        nombre = str(ws.cell(r, headers["nombre"]).value or "").strip()
        unidad = str(ws.cell(r, headers["unidad"]).value or "").strip() or "unidad"
        precio_bs = _to_float(ws.cell(r, headers["precio_bs"]).value, default=0.0)

        if not nombre and not sku:
            continue

        products.append(Product(sku=sku, categoria=categoria, nombre=nombre, unidad=unidad, precio_bs=precio_bs, activo=1))

    products.sort(key=lambda p: (p.categoria.lower(), p.nombre.lower()))
    return products


# =========================
# DB schema
# =========================
def init_db() -> None:
    if IS_POSTGRES:
        # Postgres
        db_exec("""
            CREATE TABLE IF NOT EXISTS quotes (
                id SERIAL PRIMARY KEY,
                quote_no INTEGER NOT NULL,
                created_at TIMESTAMP NOT NULL,
                client_name TEXT NOT NULL,
                delivery_time TEXT NOT NULL,
                validity_days INTEGER NOT NULL,
                notes TEXT
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS quote_items (
                id SERIAL PRIMARY KEY,
                quote_id INTEGER NOT NULL REFERENCES quotes(id) ON DELETE CASCADE,
                sku TEXT,
                name TEXT NOT NULL,
                unit TEXT NOT NULL,
                qty DOUBLE PRECISION NOT NULL,
                unit_price DOUBLE PRECISION NOT NULL
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS counter (
                id SERIAL PRIMARY KEY,
                last_quote_no INTEGER NOT NULL DEFAULT 0
            )
        """)
        db_exec("""
            INSERT INTO counter(last_quote_no)
            SELECT 0 WHERE NOT EXISTS (SELECT 1 FROM counter)
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                sku TEXT UNIQUE,
                categoria TEXT NOT NULL DEFAULT '',
                nombre TEXT NOT NULL,
                unidad TEXT NOT NULL DEFAULT 'unidad',
                precio_bs DOUBLE PRECISION NOT NULL DEFAULT 0,
                activo INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
    else:
        # SQLite
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS quotes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quote_no INTEGER NOT NULL,
                    created_at TEXT NOT NULL,
                    client_name TEXT NOT NULL,
                    delivery_time TEXT NOT NULL,
                    validity_days INTEGER NOT NULL,
                    notes TEXT
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS quote_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    quote_id INTEGER NOT NULL,
                    sku TEXT,
                    name TEXT NOT NULL,
                    unit TEXT NOT NULL,
                    qty REAL NOT NULL,
                    unit_price REAL NOT NULL,
                    FOREIGN KEY (quote_id) REFERENCES quotes(id)
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS counter (
                    key TEXT PRIMARY KEY,
                    value INTEGER NOT NULL
                )
            """)
            cur.execute("INSERT OR IGNORE INTO counter(key, value) VALUES('quote_no', 0)")
            cur.execute("""
                CREATE TABLE IF NOT EXISTS products (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sku TEXT UNIQUE,
                    categoria TEXT NOT NULL DEFAULT '',
                    nombre TEXT NOT NULL,
                    unidad TEXT NOT NULL DEFAULT 'unidad',
                    precio_bs REAL NOT NULL DEFAULT 0,
                    activo INTEGER NOT NULL DEFAULT 1,
                    created_at TEXT NOT NULL
                )
            """)
            con.commit()
        finally:
            con.close()

def products_count() -> int:
    row = db_fetchone("SELECT COUNT(*) AS total FROM products")
    return int(row["total"]) if row else 0

def seed_products_from_excel_if_empty():
    if products_count() > 0:
        return
    if not os.path.exists(EXCEL_PATH):
        return

    try:
        products = read_products_from_excel()
    except Exception:
        return

    if not products:
        return

    if IS_POSTGRES:
        for p in products:
            db_exec(
                psql("""INSERT INTO products(sku,categoria,nombre,unidad,precio_bs,activo)
                        VALUES(?,?,?,?,?,?)
                        ON CONFLICT (sku) DO UPDATE SET
                          categoria=EXCLUDED.categoria,
                          nombre=EXCLUDED.nombre,
                          unidad=EXCLUDED.unidad,
                          precio_bs=EXCLUDED.precio_bs,
                          activo=EXCLUDED.activo
                """),
                (p.sku or None, p.categoria, p.nombre, p.unidad, p.precio_bs, p.activo),
            )
    else:
        con = db_connect()
        try:
            cur = con.cursor()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for p in products:
                cur.execute("""
                    INSERT OR REPLACE INTO products(sku,categoria,nombre,unidad,precio_bs,activo,created_at)
                    VALUES(?,?,?,?,?,?,?)
                """, (p.sku or None, p.categoria, p.nombre, p.unidad, p.precio_bs, p.activo, now))
            con.commit()
        finally:
            con.close()

def load_products() -> List[Product]:
    rows = db_fetchall("""
        SELECT sku, categoria, nombre, unidad, precio_bs, activo
        FROM products
        WHERE activo=1
        ORDER BY categoria, nombre
    """)
    products: List[Product] = []
    for r in rows:
        products.append(Product(
            sku=(r["sku"] or "").strip() if r["sku"] else "",
            categoria=str(r["categoria"] or ""),
            nombre=str(r["nombre"] or ""),
            unidad=str(r["unidad"] or "unidad"),
            precio_bs=float(r["precio_bs"] or 0),
        ))
    return products


init_db()
seed_products_from_excel_if_empty()


def next_quote_no() -> int:
    if IS_POSTGRES:
        db_exec("UPDATE counter SET last_quote_no = last_quote_no + 1")
        row = db_fetchone("SELECT last_quote_no FROM counter LIMIT 1")
        return int(row["last_quote_no"])
    else:
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute("UPDATE counter SET value = value + 1 WHERE key='quote_no'")
            cur.execute("SELECT value FROM counter WHERE key='quote_no'")
            n = cur.fetchone()[0]
            con.commit()
            return int(n)
        finally:
            con.close()


# =========================
# PDF generator
# =========================
def money(x: float) -> str:
    return f"Bs {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def generate_pdf(
    quote_no: int,
    created_at: str,
    client_name: str,
    delivery_time: str,
    validity_days: int,
    items: List[dict],
    notes: Optional[str] = None,
) -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    x0 = 18 * mm
    y = height - 18 * mm

    # ---- Encabezado: texto izquierda, logo derecha ----
    logo_w = 28 * mm
    logo_h = 28 * mm
    logo_x = width - x0 - logo_w
    logo_y = y - logo_h + 4 * mm

    if LOGO_PATH and os.path.exists(LOGO_PATH):
        try:
            c.drawImage(LOGO_PATH, logo_x, logo_y, width=logo_w, height=logo_h,
                        mask="auto", preserveAspectRatio=True)
        except Exception:
            pass

    # Nombre empresa y telefono a la izquierda
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x0, y, EMPRESA_NOMBRE)
    y -= 7 * mm
    c.setFont("Helvetica", 10)
    c.drawString(x0, y, f"Telf.: {EMPRESA_TELF}")
    y -= 7 * mm

    # Linea separadora
    c.setLineWidth(0.8)
    c.line(x0, y, width - x0, y)
    y -= 8 * mm

    c.setFont("Helvetica-Bold", 13)
    c.drawString(x0, y, "COTIZACION")
    y -= 8 * mm

    c.setFont("Helvetica", 10)
    c.drawString(x0, y, f"N°: {quote_no:06d}")
    c.drawRightString(width - x0, y, f"Fecha: {created_at}")
    y -= 6 * mm
    c.drawString(x0, y, f"Cliente: {client_name}")
    y -= 6 * mm
    c.drawString(x0, y, f"Tiempo de entrega: {delivery_time}")
    y -= 6 * mm
    c.drawString(x0, y, f"Validez de la propuesta: {validity_days} día(s)")
    y -= 10 * mm

    c.setFont("Helvetica-Bold", 10)
    c.drawString(x0, y, "Ítem")
    c.drawString(x0 + 90 * mm, y, "Cant.")
    c.drawString(x0 + 110 * mm, y, "P. Unit.")
    c.drawString(x0 + 145 * mm, y, "Importe")
    y -= 4 * mm
    c.line(x0, y, width - x0, y)
    y -= 6 * mm

    c.setFont("Helvetica", 10)
    subtotal = 0.0

    for it in items:
        name = str(it["name"])
        qty = float(it["qty"])
        unit = str(it["unit"])
        unit_price = float(it["unit_price"])
        line_total = qty * unit_price
        subtotal += line_total

        shown = name if len(name) <= 52 else name[:49] + "..."
        c.drawString(x0, y, f"{shown} ({unit})")
        c.drawRightString(x0 + 105 * mm, y, f"{qty:g}")
        c.drawRightString(x0 + 140 * mm, y, money(unit_price))
        c.drawRightString(width - x0, y, money(line_total))
        y -= 6 * mm

        if y < 30 * mm:
            c.showPage()
            y = height - 18 * mm
            c.setFont("Helvetica", 10)

    y -= 2 * mm
    c.line(x0, y, width - x0, y)
    y -= 8 * mm

    iva = subtotal * IVA_RATE
    total = subtotal + iva

    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(x0 + 140 * mm, y, "Subtotal (Sin IVA):")
    c.drawRightString(width - x0, y, money(subtotal))
    y -= 6 * mm

    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(x0 + 140 * mm, y, f"IVA ({int(IVA_RATE * 100)}%):")
    c.drawRightString(width - x0, y, money(iva))
    y -= 6 * mm

    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(x0 + 140 * mm, y, "Total (Con IVA):")
    c.drawRightString(width - x0, y, money(total))
    y -= 10 * mm

    if notes:
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x0, y, "Notas:")
        y -= 6 * mm
        c.setFont("Helvetica", 10)
        for line in str(notes).splitlines():
            c.drawString(x0, y, line[:95])
            y -= 5 * mm

    c.setFont("Helvetica", 8)
    c.drawString(x0, 12 * mm, f"{EMPRESA_NOMBRE} - Cotización generada automáticamente")
    c.showPage()
    c.save()
    return buf.getvalue()


# =========================
# Auth routes
# =========================
@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request):
    if is_logged_in(request):
        return RedirectResponse(url="/nueva", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "err": request.query_params.get("err")})

@app.post("/login")
def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    if username == ADMIN_USER and password == ADMIN_PASS:
        request.session["auth"] = True
        return RedirectResponse(url="/nueva", status_code=303)
    return RedirectResponse(url="/login?err=Usuario+o+clave+incorrecta", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)


# =========================
# Routes
# =========================
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    if not is_logged_in(request):
        return RedirectResponse(url="/login", status_code=303)
    return RedirectResponse(url="/dashboard", status_code=303)

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    gate = require_login(request)
    if gate:
        return gate

    from calendar import month_name
    hoy = datetime.now()
    hoy_str = hoy.strftime("%Y-%m-%d")
    mes_actual = hoy.strftime("%Y-%m")

    # Stats instalaciones
    if IS_POSTGRES:
        row_inst = db_fetchone("""
            SELECT
              COUNT(*) AS total,
              SUM(CASE WHEN estado='pendiente' THEN 1 ELSE 0 END) AS pendientes,
              SUM(CASE WHEN estado='en_curso' THEN 1 ELSE 0 END) AS en_curso,
              SUM(CASE WHEN estado='completada' THEN 1 ELSE 0 END) AS completadas
            FROM instalaciones
        """)
    else:
        row_inst = db_fetchone("""
            SELECT
              COUNT(*) AS total,
              SUM(CASE WHEN estado='pendiente' THEN 1 ELSE 0 END) AS pendientes,
              SUM(CASE WHEN estado='en_curso' THEN 1 ELSE 0 END) AS en_curso,
              SUM(CASE WHEN estado='completada' THEN 1 ELSE 0 END) AS completadas
            FROM instalaciones
        """)

    total_inst = int(row_inst["total"] or 0) if row_inst else 0
    pendientes = int(row_inst["pendientes"] or 0) if row_inst else 0
    en_curso = int(row_inst["en_curso"] or 0) if row_inst else 0
    completadas = int(row_inst["completadas"] or 0) if row_inst else 0

    # Ventas del mes actual (cotizaciones del mes)
    if IS_POSTGRES:
        rows_mes = db_fetchall(
            "SELECT id FROM quotes WHERE created_at >= %s AND created_at < %s",
            (f"{mes_actual}-01", f"{hoy.year}-{hoy.month+1:02d}-01" if hoy.month < 12 else f"{hoy.year+1}-01-01"),
        )
    else:
        rows_mes = db_fetchall(
            "SELECT id FROM quotes WHERE substr(created_at,1,7) = ?", (mes_actual,)
        )

    ventas_mes = sum(get_quote_total(int(r["id"])) for r in rows_mes)
    ventas_mes_sin_iva = ventas_mes / (1 + IVA_RATE)

    # Próximas instalaciones (7 días)
    fecha_limite = (hoy + __import__('datetime').timedelta(days=7)).strftime("%Y-%m-%d")
    if IS_POSTGRES:
        proximas_rows = db_fetchall("""
            SELECT i.estado, i.tecnico, i.fecha_instalacion, q.client_name, q.id as quote_id
            FROM instalaciones i JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion BETWEEN %s AND %s
            ORDER BY i.fecha_instalacion
            LIMIT 8
        """, (hoy_str, fecha_limite))
    else:
        proximas_rows = db_fetchall("""
            SELECT i.estado, i.tecnico, i.fecha_instalacion, q.client_name, q.id as quote_id
            FROM instalaciones i JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion BETWEEN ? AND ?
            ORDER BY i.fecha_instalacion
            LIMIT 8
        """, (hoy_str, fecha_limite))

    proximas = [dict(r) for r in proximas_rows]

    # Cotizaciones recientes
    if IS_POSTGRES:
        cots_rows = db_fetchall("SELECT id, quote_no, client_name, created_at FROM quotes ORDER BY id DESC LIMIT 6")
    else:
        cots_rows = db_fetchall("SELECT id, quote_no, client_name, created_at FROM quotes ORDER BY id DESC LIMIT 6")

    @dataclass
    class CotRow:
        id: int
        quote_no: int
        client_name: str
        created_at_str: str

    cotizaciones_recientes = []
    for r in cots_rows:
        ca = r["created_at"]
        ca_str = ca.strftime("%d/%m/%Y") if hasattr(ca, "strftime") else str(ca)[:10]
        cotizaciones_recientes.append(CotRow(
            id=int(r["id"]), quote_no=int(r["quote_no"]),
            client_name=str(r["client_name"]), created_at_str=ca_str,
        ))

    # Ventas últimos 6 meses para el gráfico
    import datetime as dt
    meses_labels = []
    meses_valores = []
    for i in range(5, -1, -1):
        d = hoy - dt.timedelta(days=30 * i)
        ym = d.strftime("%Y-%m")
        label = d.strftime("%b %Y")
        if IS_POSTGRES:
            primer_dia = dt.date(d.year, d.month, 1)
            if d.month == 12:
                ultimo_dia = dt.date(d.year + 1, 1, 1)
            else:
                ultimo_dia = dt.date(d.year, d.month + 1, 1)
            rows_m = db_fetchall(
                "SELECT id FROM quotes WHERE created_at >= %s AND created_at < %s",
                (str(primer_dia), str(ultimo_dia)),
            )
        else:
            rows_m = db_fetchall("SELECT id FROM quotes WHERE substr(created_at,1,7) = ?", (ym,))
        total_m = sum(get_quote_total(int(r["id"])) for r in rows_m)
        meses_labels.append(label)
        meses_valores.append(round(total_m, 2))

    meses_nombres = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                     "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes_nombre = meses_nombres[hoy.month - 1]

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "empresa": EMPRESA_NOMBRE,
        "hoy": hoy.strftime("%d/%m/%Y"),
        "mes_nombre": mes_nombre,
        "stats": {
            "pendientes": pendientes,
            "en_curso": en_curso,
            "completadas": completadas,
            "total_instalaciones": total_inst,
            "ventas_mes_con_iva": ventas_mes,
            "ventas_mes_sin_iva": ventas_mes_sin_iva,
            "cotizaciones_mes": len(rows_mes),
        },
        "proximas_instalaciones": proximas,
        "cotizaciones_recientes": cotizaciones_recientes,
        "meses_labels": meses_labels,
        "meses_valores": meses_valores,
    })

@app.get("/nueva", response_class=HTMLResponse)
def nueva(request: Request):
    gate = require_login(request)
    if gate:
        return gate
    products = load_products()
    return templates.TemplateResponse(
        "nueva.html",
        {"request": request, "products": products, "empresa": EMPRESA_NOMBRE, "telf": EMPRESA_TELF, "iva_rate": IVA_RATE},
    )

@app.post("/crear")
def crear_cotizacion(
    request: Request,
    client_name: str = Form(...),
    delivery_time: str = Form(...),
    validity_days: int = Form(...),
    notes: str = Form(""),
    item_sku: List[str] = Form([]),
    item_name: List[str] = Form([]),
    item_unit: List[str] = Form([]),
    item_qty: List[float] = Form([]),
    item_unit_price: List[float] = Form([]),
):
    gate = require_login(request)
    if gate:
        return gate

    items = []
    for i in range(len(item_name)):
        name = (item_name[i] or "").strip()
        if not name:
            continue
        qty = float(item_qty[i] or 0)
        price = float(item_unit_price[i] or 0)
        if qty <= 0:
            continue
        items.append({
            "sku": (item_sku[i] or "").strip(),
            "name": name,
            "unit": (item_unit[i] or "unidad").strip(),
            "qty": qty,
            "unit_price": price,
        })

    if not items:
        return RedirectResponse(url="/nueva?err=Agrega+al+menos+un+item", status_code=303)

    qno = next_quote_no()

    if IS_POSTGRES:
        created_at = datetime.now()
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute(
                psql("INSERT INTO quotes(quote_no, created_at, client_name, delivery_time, validity_days, notes) VALUES(?,?,?,?,?,?) RETURNING id"),
                (qno, created_at, client_name.strip(), delivery_time.strip(), int(validity_days), notes.strip() or None),
            )
            quote_id = int(cur.fetchone()["id"])

            for it in items:
                cur.execute(
                    psql("INSERT INTO quote_items(quote_id, sku, name, unit, qty, unit_price) VALUES(?,?,?,?,?,?)"),
                    (quote_id, it["sku"] or None, it["name"], it["unit"], it["qty"], it["unit_price"]),
                )

            con.commit()
        finally:
            con.close()
    else:
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute(
                "INSERT INTO quotes(quote_no, created_at, client_name, delivery_time, validity_days, notes) VALUES(?,?,?,?,?,?)",
                (qno, created_at, client_name.strip(), delivery_time.strip(), int(validity_days), notes.strip() or None),
            )
            quote_id = cur.lastrowid

            for it in items:
                cur.execute(
                    "INSERT INTO quote_items(quote_id, sku, name, unit, qty, unit_price) VALUES(?,?,?,?,?,?)",
                    (quote_id, it["sku"], it["name"], it["unit"], it["qty"], it["unit_price"]),
                )
            con.commit()
        finally:
            con.close()

    return RedirectResponse(url=f"/cotizacion/{quote_id}/pdf", status_code=303)

@app.get("/historial", response_class=HTMLResponse)
def historial(request: Request):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        rows = db_fetchall("SELECT id, quote_no, created_at, client_name, delivery_time, validity_days, notes FROM quotes ORDER BY id DESC LIMIT 500")
    else:
        rows = db_fetchall("SELECT * FROM quotes ORDER BY id DESC LIMIT 500")

    return templates.TemplateResponse("historial.html", {"request": request, "quotes": rows, "empresa": EMPRESA_NOMBRE})

@app.get("/cotizacion/{quote_id}/pdf")
def cotizacion_pdf(request: Request, quote_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        q = db_fetchone("SELECT * FROM quotes WHERE id=%s", (quote_id,))
        if not q:
            return RedirectResponse(url="/historial", status_code=303)

        items_rows = db_fetchall(
            "SELECT sku, name, unit, qty, unit_price FROM quote_items WHERE quote_id=%s ORDER BY id",
            (quote_id,),
        )

        items = []
        for r in items_rows:
            items.append({
                "sku": (r["sku"] or ""),
                "name": r["name"],
                "unit": r["unit"],
                "qty": float(r["qty"]),
                "unit_price": float(r["unit_price"]),
            })

        created_at = q["created_at"].strftime("%Y-%m-%d %H:%M") if hasattr(q["created_at"], "strftime") else str(q["created_at"])

        pdf = generate_pdf(
            quote_no=int(q["quote_no"]),
            created_at=created_at,
            client_name=str(q["client_name"]),
            delivery_time=str(q["delivery_time"]),
            validity_days=int(q["validity_days"]),
            items=items,
            notes=q["notes"],
        )

        filename = f"cotizacion_{int(q['quote_no']):06d}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="{filename}"'},
        )

    # --- SQLite (tu código actual) ---
    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute("SELECT * FROM quotes WHERE id=?", (quote_id,))
        q = cur.fetchone()
        if not q:
            return RedirectResponse(url="/historial", status_code=303)

        cur.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY id", (quote_id,))
        items = [dict(r) for r in cur.fetchall()]
    finally:
        con.close()

    pdf = generate_pdf(
        quote_no=int(q["quote_no"]),
        created_at=str(q["created_at"]),
        client_name=str(q["client_name"]),
        delivery_time=str(q["delivery_time"]),
        validity_days=int(q["validity_days"]),
        items=items,
        notes=q["notes"],
    )

    filename = f"cotizacion_{int(q['quote_no']):06d}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )

@app.get("/cotizacion/{quote_id}/editar", response_class=HTMLResponse)
def editar_get(request: Request, quote_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        q = db_fetchone("SELECT * FROM quotes WHERE id=%s", (quote_id,))
        if not q:
            return RedirectResponse(url="/historial", status_code=303)

        items_rows = db_fetchall("SELECT id, sku, name, unit, qty, unit_price FROM quote_items WHERE quote_id=%s ORDER BY id", (quote_id,))
        return templates.TemplateResponse("editar.html", {
            "request": request,
            "q": q,
            "items": items_rows,
            "products": load_products(),
            "empresa": EMPRESA_NOMBRE,
            "telf": EMPRESA_TELF,
            "iva_rate": IVA_RATE,
        })

    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute("SELECT * FROM quotes WHERE id=?", (quote_id,))
        q = cur.fetchone()
        cur.execute("SELECT * FROM quote_items WHERE quote_id=? ORDER BY id", (quote_id,))
        items = cur.fetchall()
    finally:
        con.close()

    if not q:
        return RedirectResponse(url="/historial", status_code=303)

    return templates.TemplateResponse("editar.html", {
        "request": request,
        "q": q,
        "items": items,
        "products": load_products(),
        "empresa": EMPRESA_NOMBRE,
        "telf": EMPRESA_TELF,
        "iva_rate": IVA_RATE,
    })

@app.post("/cotizacion/{quote_id}/editar")
def editar_post(
    request: Request,
    quote_id: int,
    client_name: str = Form(...),
    delivery_time: str = Form(...),
    validity_days: int = Form(...),
    notes: str = Form(""),
    item_sku: List[str] = Form([]),
    item_name: List[str] = Form([]),
    item_unit: List[str] = Form([]),
    item_qty: List[float] = Form([]),
    item_unit_price: List[float] = Form([]),
):
    gate = require_login(request)
    if gate:
        return gate

    items = []
    for i in range(len(item_name)):
        name = (item_name[i] or "").strip()
        if not name:
            continue
        qty = float(item_qty[i] or 0)
        price = float(item_unit_price[i] or 0)
        if qty <= 0:
            continue
        items.append({
            "sku": (item_sku[i] or "").strip(),
            "name": name,
            "unit": (item_unit[i] or "unidad").strip(),
            "qty": qty,
            "unit_price": price,
        })

    if not items:
        return RedirectResponse(url=f"/cotizacion/{quote_id}/editar?err=Agrega+al+menos+un+item", status_code=303)

    if IS_POSTGRES:
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute(
                psql("""UPDATE quotes SET client_name=?, delivery_time=?, validity_days=?, notes=? WHERE id=?"""),
                (client_name.strip(), delivery_time.strip(), int(validity_days), notes.strip() or None, quote_id),
            )
            cur.execute("DELETE FROM quote_items WHERE quote_id=%s", (quote_id,))
            for it in items:
                cur.execute(
                    psql("INSERT INTO quote_items(quote_id, sku, name, unit, qty, unit_price) VALUES(?,?,?,?,?,?)"),
                    (quote_id, it["sku"] or None, it["name"], it["unit"], it["qty"], it["unit_price"]),
                )
            con.commit()
        finally:
            con.close()
        return RedirectResponse(url="/historial", status_code=303)

    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute("""
            UPDATE quotes
            SET client_name=?, delivery_time=?, validity_days=?, notes=?
            WHERE id=?
        """, (client_name.strip(), delivery_time.strip(), int(validity_days), notes.strip() or None, quote_id))

        cur.execute("DELETE FROM quote_items WHERE quote_id=?", (quote_id,))
        for it in items:
            cur.execute(
                "INSERT INTO quote_items(quote_id, sku, name, unit, qty, unit_price) VALUES(?,?,?,?,?,?)",
                (quote_id, it["sku"], it["name"], it["unit"], it["qty"], it["unit_price"]),
            )
        con.commit()
    finally:
        con.close()

    return RedirectResponse(url="/historial", status_code=303)

@app.post("/cotizacion/{quote_id}/borrar")
def borrar(request: Request, quote_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute("DELETE FROM quotes WHERE id=%s", (quote_id,))
            con.commit()
        finally:
            con.close()
        return RedirectResponse(url="/historial", status_code=303)

    con = db_connect()
    try:
        cur = con.cursor()
        cur.execute("DELETE FROM quote_items WHERE quote_id=?", (quote_id,))
        cur.execute("DELETE FROM quotes WHERE id=?", (quote_id,))
        con.commit()
    finally:
        con.close()
    return RedirectResponse(url="/historial", status_code=303)


# =========================
# Gestión de productos
# =========================
@dataclass
class ProductoRow:
    id: int
    sku: str
    categoria: str
    nombre: str
    unidad: str
    precio_bs: float
    activo: int

def load_all_products() -> List[ProductoRow]:
    if IS_POSTGRES:
        rows = db_fetchall("SELECT id, sku, categoria, nombre, unidad, precio_bs, activo FROM products ORDER BY categoria, nombre")
    else:
        rows = db_fetchall("SELECT id, sku, categoria, nombre, unidad, precio_bs, activo FROM products ORDER BY categoria, nombre")
    result = []
    for r in rows:
        result.append(ProductoRow(
            id=int(r["id"]),
            sku=str(r["sku"] or ""),
            categoria=str(r["categoria"] or ""),
            nombre=str(r["nombre"] or ""),
            unidad=str(r["unidad"] or "unidad"),
            precio_bs=float(r["precio_bs"] or 0),
            activo=int(r["activo"]),
        ))
    return result

@app.get("/productos", response_class=HTMLResponse)
def productos_get(request: Request, msg: str = "", msg_type: str = "success"):
    gate = require_login(request)
    if gate:
        return gate
    productos = load_all_products()
    return templates.TemplateResponse("productos.html", {
        "request": request,
        "productos": productos,
        "empresa": EMPRESA_NOMBRE,
        "msg": request.query_params.get("msg", ""),
        "msg_type": request.query_params.get("msg_type", "success"),
    })

@app.post("/productos/guardar")
def productos_guardar(
    request: Request,
    producto_id: str = Form(""),
    sku: str = Form(""),
    categoria: str = Form(...),
    nombre: str = Form(...),
    unidad: str = Form(...),
    precio_bs: float = Form(...),
):
    gate = require_login(request)
    if gate:
        return gate

    sku = sku.strip() or None
    categoria = categoria.strip()
    nombre = nombre.strip()
    unidad = unidad.strip()

    if producto_id:
        # Editar existente
        if IS_POSTGRES:
            db_exec(
                "UPDATE products SET sku=%s, categoria=%s, nombre=%s, unidad=%s, precio_bs=%s WHERE id=%s",
                (sku, categoria, nombre, unidad, precio_bs, int(producto_id)),
            )
        else:
            db_exec(
                "UPDATE products SET sku=?, categoria=?, nombre=?, unidad=?, precio_bs=? WHERE id=?",
                (sku, categoria, nombre, unidad, precio_bs, int(producto_id)),
            )
        msg = "Producto actualizado correctamente."
    else:
        # Nuevo producto
        if IS_POSTGRES:
            db_exec(
                "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo) VALUES(%s,%s,%s,%s,%s,1)",
                (sku, categoria, nombre, unidad, precio_bs),
            )
        else:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db_exec(
                "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo, created_at) VALUES(?,?,?,?,?,1,?)",
                (sku, categoria, nombre, unidad, precio_bs, now),
            )
        msg = "Producto agregado correctamente."

    return RedirectResponse(url=f"/productos?msg={msg}&msg_type=success", status_code=303)

@app.post("/productos/{producto_id}/toggle")
def productos_toggle(request: Request, producto_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        db_exec("UPDATE products SET activo = CASE WHEN activo=1 THEN 0 ELSE 1 END WHERE id=%s", (producto_id,))
    else:
        db_exec("UPDATE products SET activo = CASE WHEN activo=1 THEN 0 ELSE 1 END WHERE id=?", (producto_id,))

    return RedirectResponse(url="/productos?msg=Estado+actualizado.&msg_type=success", status_code=303)

@app.post("/productos/{producto_id}/borrar")
def productos_borrar(request: Request, producto_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        db_exec("DELETE FROM products WHERE id=%s", (producto_id,))
    else:
        db_exec("DELETE FROM products WHERE id=?", (producto_id,))

    return RedirectResponse(url="/productos?msg=Producto+eliminado.&msg_type=warning", status_code=303)


# =========================
# Carga masiva de productos
# =========================
CAMARAS_WIFI = [
    ("CL1B-5-E27","IMOU - Smart","Bombilla Smart Wifi LED 2.4G - Control Grupo - Colores - Alexa/Google - 70M","unidad",111),
    ("IPC-C22FN-C-IMOU","IMOU - Cámaras Wifi","Cámara Wifi Versa 2MP Full Color - Lente 2.8mm - 114 - H.265 - Audio 2 Vías - Sirena - MicroSD 256GB - IP65","unidad",455),
    ("IPC-GK2CN-3C1WR-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Ranger RC 3MP - Rastreo Inteligente - Modo Privacidad - Llamada 1 Toque - Sirena - MicroSD 256GB - Audio Bidireccional","unidad",361),
    ("IPC-GS7EN-3M0WE-IMOU","IMOU - Cámaras Wifi","Cámara Cruiser 2 IP Wifi 2K - Detec. Personas/Vehículos - Full Color - IP66 - Audio Bidireccional - Rastreo - MicroSD 256GB - Lente 3.6mm","unidad",674),
    ("IPC-GS7EN-5M0WE-IMOU","IMOU - Cámaras Wifi","Cámara Cruiser IP Wifi 5MP Full Color - Detec. Humanos/Vehículos - Disuasión Activa - IP66 - Audio Bidireccional - MicroSD 256GB - Rastreo","unidad",728),
    ("IPC-K2EN-3H3W-IMOU","IMOU - Cámaras Wifi","Cámara Ranger 2 IP Wifi 3MP - Audio Bidireccional - Auto Crucero - Sirena - MicroSD 512GB - Rastreo Inteligente - ONVIF","unidad",312),
    ("IPC-K2EN-5H3W-IMOU","IMOU - Cámaras Wifi","Cámara Ranger 2 IP Wifi 5MP - Audio Bidireccional - Auto Crucero - Sirena - MicroSD 512GB - Rastreo Inteligente - ONVIF","unidad",440),
    ("IPC-K3DN-3H0WF-0280B-IMOU","IMOU - Cámaras Wifi","Cámara Bala Bullet 2E 3MP Full Color - Detec. Humanos - IP67 - Mic Integrado - MicroSD 512GB - RJ45 - Lente 2.8mm","unidad",369),
    ("IPC-K3DN-5H0WF-0280B-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Bala 5MP Bullet 2E Full Color - Plástica - Lente 2.8mm - Detec. Humanos - Mic Integrado - MicroSD 512GB - IP67","unidad",449),
    ("IPC-K7FN-3H0WE-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Cruiser SC 3MP - Tipo PT - Full Color - Rastreo Inteligente - Audio Bidireccional - RJ45 - MicroSD 512GB - Exterior - Disuasiva","unidad",465),
    ("IPC-K7FN-5H0WE-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Cruiser SC+ 5MP - Detec. Humanos - Rastreo Inteligente - Disuasión Activa - IP66 - Audio Bidireccional - SD 512GB - Full Color","unidad",517),
    ("IPC-K9DCN-3T0WE/FSP12","IMOU - Kits Solar","Kit IP Wifi Cell 3C Exterior 3MP - Batería Integrada - Sirena/Reflector - Audio Bidireccional - IP66 - MicroSD 256GB - Panel Solar","unidad",863),
    ("IPC-K9ECN-3T0WE/FSP12","IMOU - Kits Solar","Kit IP Wifi Cell PT Lite 3MP - Batería Integrada - Sirena/Reflector - Audio Bidireccional - IP66 - MicroSD 256GB - Panel Solar","unidad",1124),
    ("IPC-S2EN-3R1S-IMOU","IMOU - Cámaras Wifi","Cámara Ranger 2 IP Wifi 3MP - Llamada 1 Toque - Detec. Personas/Mascotas - Audio Bidireccional - Full Color - MicroSD 512GB","unidad",335),
    ("IPC-S2EN-5R1S-IMOU","IMOU - Cámaras Wifi","Cámara Ranger 2 IP Wifi 5MP - Llamada 1 Toque - Detec. Personas/Mascotas - Audio Bidireccional - Full Color - MicroSD 512GB","unidad",399),
    ("IPC-S3EN-3M0WE-0280B-IMOU","IMOU - Cámaras Wifi","Cámara IP Inalámbrica Bala 3MP Full Color - Detec. Humanos/Vehículos - IP67 - Audio Doble Vía - POE - MicroSD 256GB - Plástica/Metal","unidad",480),
    ("IPC-S3EN-5M0WE-0280B-IMOU","IMOU - Cámaras Wifi","Cámara Wifi Bullet 3 5MP Full Color - Detec. Humanos/Vehículos - IP67 - Audio Doble Vía - MicroSD 256GB - Plástica/Metal","unidad",567),
    ("IPC-S2XN-6M0WED-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Ranger Dual 6MP - Doble Lente - Detec. IA Personas/Vehículos - Disuasión Activa - Audio Bidireccional - Full Color - MicroSD 256GB","unidad",515),
    ("IPC-S2XEN-6M0S-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Ranger Dual 3MP+3MP - Botón Llamada - Detec. Personas/Mascotas - Rastreo Inteligente - Audio Bidireccional - MicroSD 512GB","unidad",545),
    ("IPC-S6DN-3M0WEB-E27-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Tipo Foco 3MP - PTZ - Disuasión Activa - Detec. IA Vehículos/Personas - Full Color - Audio Bidireccional - MicroSD 256GB - ONVIF","unidad",380),
    ("IPC-T26EN-0280B-IMOU","IMOU - Cámaras Wifi","Cámara IP Turret Wifi 2MP - Lente 2.8mm - IR 30M - H.265/H.264 - Audio Doble Vía - Sirena/Reflector - ONVIF - MicroSD 256GB - IP67","unidad",597),
    ("IPC-S7XEN-8M0WED-0360B","IMOU - Cámaras Wifi","Cámara Cruiser Dual 2 IP Wifi 3+5MP - Tipo PT - Luces Disuasivas Rojo/Azul - Sirena - Full Color - Audio Bidireccional","unidad",699),
    ("IPC-S7UN-11M0WED-IMOU","IMOU - Cámaras Wifi","Cámara IP Wifi Cruiser Triple 3+3+5MP - 2 Ranuras MicroSD 1024GB - Audio Bidireccional - Visión Nocturna 30M - Exterior","unidad",961),
    ("DH-IPC-C3AP-0280B","Dahua - Cámaras Wifi","Cámara IP Cubo Wifi 3MP - Interior - Detec. Humanos IA - Detec. Mascotas - Audio Bidireccional - Wifi 6 - Bluetooth - MicroSD 256GB","unidad",283),
    ("DH-IPC-C5AP-0280B","Dahua - Cámaras Wifi","Cámara IP Wifi Cube C1 5MP - Interior - Detec. Humanos IA - Detec. Mascotas - Audio Bidireccional - Wifi 6 - Bluetooth - MicroSD 256GB","unidad",335),
    ("DH-IPC-H3AP-0360B","Dahua - Cámaras Wifi","Cámara IP Hero A1 Wifi 3MP Tipo PT - 360 - Audio Bidireccional - Detec. IA Humanos - Detec. Mascotas - Wifi 6 - Auto Tracking - MicroSD 256GB","unidad",357),
    ("DH-IPC-H5AP-0360B","Dahua - Cámaras Wifi","Cámara IP Hero 1 Wifi 5MP Tipo PT - 360 - Audio Bidireccional - Detec. IA Humanos - Detec. Mascotas - Wifi 6 - Auto Tracking - MicroSD 256GB","unidad",389),
    ("DH-IPC-H3BP-0360B","Dahua - Cámaras Wifi","Cámara IP Wifi 3MP - Llamada 1 Toque - Rotación 360 - Audio Bidireccional - Detec. IA Humanos - Detec. Mascotas - Rastreo Inteligente","unidad",364),
    ("DH-IPC-H5BP-0360B","Dahua - Cámaras Wifi","Cámara Hero B1 IP Wifi 5MP - Rotación 360 - Llamada 1 Toque - Audio Bidireccional - Detec. IA Humanos - Detec. Mascotas - Pareo Bluetooth","unidad",424),
    ("DH-SD-H4C-0400B","Dahua - Cámaras Wifi","Cámara IP Wifi 4MP Tipo PT - Sensor CMOS - 25/30fps@1080p - Detec. Personas - Audio Bidireccional - Alarma Integrada - RJ45","unidad",494),
    ("DH-IPC-H3DP-3F-0360B","Dahua - Cámaras Wifi","Cámara Wifi Tipo PT Doble Lente 3+3MP - Ángulos Ajustables - Llamada 1 Toque - Detec. Personas/Mascotas - Auto Tracking - MicroSD 256GB","unidad",520),
    ("DH-IPC-H5DP-5F-0360B","Dahua - Cámaras Wifi","Cámara Wifi Tipo PT Doble Lente 5+5MP - Ángulos Ajustables - Llamada 1 Toque - Detec. Personas/Mascotas - Auto Tracking - MicroSD 256GB","unidad",604),
    ("DH-IPC-P3ASP-PV-0400B-S2","Dahua - Cámaras Wifi","Cámara IP Wifi Picoo 3MP - Lente 4mm - Alarma Sonido/Luz - Detec. IA Personas/Vehículos - IP66 - Modo Patrullaje - Audio Bidireccional - MicroSD 256GB","unidad",515),
    ("DH-IPC-P5ASP-PV-0400B-S2","Dahua - Cámaras Wifi","Cámara IP Wifi 5MP Tipo PT - Detec. IA Personas/Vehículos - Audio Bidireccional - Auto Tracking - IP66 - Alarma Sonido/Luz - MicroSD 256GB","unidad",602),
    ("DH-IPC-P3BP-PV-0360B","Dahua - Cámaras Wifi","Cámara IP Picoo B1 Wifi 3MP Tipo PT - Smart Dual Light - 360 - Alarma Sonido/Luz - Detec. IA Personas/Vehículos - Wifi 6 - MicroSD 256GB","unidad",499),
    ("DH-IPC-P5BP-PV-0360B","Dahua - Cámaras Wifi","Cámara IP Wifi Picoo B1 5MP TIOC - Smart Dual Light - 360 - Detec. Humanos/Vehículos IA - Auto Tracking - Wifi 6 - IP66 - MicroSD 256GB","unidad",620),
    ("DH-IPC-P5FP-PV-0360B-PRO","Dahua - Cámaras Wifi","Cámara IP Wifi 5MP WizColor PT - Rastreo Inteligente - Detec. IA Personas/Vehículos - IP66 - Audio Bidireccional - Alarma Sonido/Luz - MicroSD 512GB","unidad",703),
    ("IPC-P3DP-3F-PV-0280B","Dahua - Cámaras Wifi","Cámara IP Wifi 3x3MP Doble Lente Tipo PT - Detección Dual - Control de Ronda - Sirena/Luz - Detec. IA Personas/Vehículos - IP66 - MicroSD 256GB","unidad",772),
    ("DH-IPC-P3DP-3F-PV-P-PRO","Dahua - Cámaras Wifi","Cámara IP Tipo PT Wifi Doble Lente 3+3MP WizColor - Audio Bidireccional - Wifi 6 - Detec. Personas/Vehículos - IP66 - MicroSD 256GB - Rastreo","unidad",1050),
    ("IPC-P5DP-5F-PV-0280B","Dahua - Cámaras Wifi","Cámara IP Wifi 5x5MP Doble Lente Tipo PT - Detección Dual - Control de Ronda - Sirena/Luz - Detec. IA Personas/Vehículos - IP66 - MicroSD 256GB","unidad",865),
    ("DH-IPC-F2CN-PV-0280B","Dahua - Cámaras Wifi","Cámara IP Wifi 2MP Tipo Bala - Detec. Humanos IA - Smart Dual Light - Disuasión Activa - Audio Bidireccional - IP67 - MicroSD 256GB","unidad",655),
    ("DH-F3D-PV","Dahua - Cámaras Wifi","Cámara Wifi Tipo Bala 3MP - IP67 - Detec. Humanos IA - MicroSD 256GB - Smart Dual Light - Disuasión Activa - Audio Bidireccional","unidad",450),
    ("DH-IPC-T2AN-PV-0280B","Dahua - Cámaras Wifi","Cámara IP Wifi Domo 2MP - Lente 2.8mm - Detec. Humanos AI - Disuasión Activa - Audio Bidireccional - IP67 - MicroSD 256GB","unidad",630),
    ("DH-P4F-PV-4G","Dahua - Cámaras 4G","Cámara Tipo PT 4MP 4G WizColor - 360 - MicroSD 256GB - Audio Bidireccional","unidad",1092),
    ("DH-IPC-HFW2441DG-4G-SP-B","Dahua - Cámaras 4G","Cámara IP Bala 4G 4MP - Panel Solar - Batería - SMD - Sirena/Luz Disuasión - Audio Bidireccional - IR 50M - Luz Cálida 30M - MicroSD - IP67","unidad",3638),
    ("DMS-BAT-L1+SOL-SP5","Smart Wifi - Exterior","Cámara Smart Exterior con Batería Recargable 6 meses - 2MP - Soporta Panel Solar DMS-SOL-SP5 - Requiere MicroSD","unidad",466),
    ("DMS-REF-D7","Smart Wifi - Exterior","Cámara Smart Exterior con Reflectores LED 2500 Lúmenes - 3MP - Compatible App Smart Life y Tuya - Requiere MicroSD","unidad",901),
    ("DMS-LB002-E26-W","Smart Wifi - Accesorios","Socket Smart para automatizar luces tipo bombillos E26","unidad",123),
    ("G103","Smart Wifi - Interior","Cámara Cubo Smart 3MP - Interior - Detec. Movimiento - Audio Bidireccional - 100 grados - MicroSD 128GB - IR 10M","unidad",149),
    ("G104","Smart Wifi - Interior","Cámara Smart PTZ 3MP - Interior - Detec. Movimiento - Audio Bidireccional - MicroSD 128GB - Visión Nocturna - Wifi 2.4GHz","unidad",189),
    ("G105","Smart Wifi - Interior","Cámara Socket PTZ 4MP - Interior - Visión Nocturna - Detec. Movimiento - Audio Bidireccional - MicroSD 128GB - Puerto E27","unidad",218),
    ("G106","Smart Wifi - Interior","Cámara Panorámica 3MP FHD - Interior - 360 grados - MicroSD 128GB - Visión Nocturna - Audio Bidireccional - Diámetro 7cm","unidad",253),
    ("G116","Smart Wifi - Interior","Cámara IP Wifi 3MP - Llamada 1 Toque - Lente 3.6mm - Visión Nocturna 10M - MicroSD 256GB - Rotación 355 grados - Rastreo - Audio Bidireccional","unidad",215),
    ("G117","Smart Wifi - Interior","Cámara para Bebé IP Wifi 3MP - Detec. Personas/Llanto - Canciones de Cuna - Recordatorio Alimentación - Audio Bidireccional - MicroSD 256GB","unidad",216),
    ("G118","Smart Wifi - Exterior","Cámara PT 4G con Panel Solar y Batería - 3MP - Full Color - Sensor PIR - Audio Bidireccional - Visión Nocturna 15M - IP65 - MicroSD 256GB","unidad",870),
    ("G119","Smart Wifi - Exterior","Cámara PT Wifi 4MP con Panel Solar y Batería - Full Color - Sensor PIR - Audio Bidireccional - Visión Nocturna 15M - IP65 - MicroSD 256GB","unidad",860),
    ("G120","Smart Wifi - Exterior","Cámara PT Wifi 4MP Tipo Foco - Detec. Personas/Movimiento - Audio Bidireccional - IR 8-10M - Wifi 2.4GHz - MicroSD 256GB","unidad",235),
    ("G121","Smart Wifi - Exterior","Cámara Wifi Fija 3MP Full Color - Detec. Personas/Movimiento - Ilum. 8M - Audio Bidireccional - Wifi 2.4GHz - MicroSD 256GB","unidad",207),
    ("G122","Smart Wifi - Exterior","Cámara PT Wifi 4MP Exterior - Rastreo Movimiento - Detec. Personas - Full Color - IP65 - Audio Bidireccional - Ilum. 10M - MicroSD 256GB","unidad",335),
    ("G123","Smart Wifi - Accesorios","Kit Timbre Inalámbrico 4MP - Batería 5200mAh - Sensor PIR 9M - IP64 - Audio Bidireccional - Alarma Antidesmontaje - Duración 2-3 meses","unidad",450),
    ("TAPO-C200","Tapo - Cámaras Wifi","Cámara Wifi Tapo C200 - Rotación 360 - Full HD 1080P - Audio 2 Vías - Visión Nocturna - Detec. Movimiento - Modo Privacidad - MicroSD 128GB","unidad",246),
    ("TAPO-C210","Tapo - Cámaras Wifi","Cámara Wifi Tapo C210 - Rotación 360 - 2K 3MP - Audio Bidireccional - Visión Nocturna - Detec. Movimiento - MicroSD 512GB","unidad",262),
    ("TAPO-C220","Tapo - Cámaras Wifi","Cámara Wifi Tapo C220 - 2K QHD 4MP - Detec. IA Inteligente - Detec. Llanto Bebé - Visión Nocturna 10M - Interior - Detec. Mascota","unidad",330),
    ("TAPO-C225","Tapo - Cámaras Wifi","Cámara Wifi Tapo C225 - 2K QHD 4MP - Detec. IA - Modo Privado Manual - Alarma Luz/Sonido - Visión Nocturna - Interior","unidad",505),
    ("TAPO-C310","Tapo - Cámaras Wifi","Cámara Wifi Outdoor Tapo C310 - 3MP - Audio Bidireccional - Visión Nocturna - Detec. Movimiento - MicroSD 128GB - IP66","unidad",427),
    ("TAPO-C320WS","Tapo - Cámaras Wifi","Cámara Wifi Outdoor Tapo C320WS - 2K QHD - Audio Bidireccional - Visión Nocturna - Detec. Movimiento - MicroSD 256GB - IP66","unidad",477),
    ("TAPO-C500","Tapo - Cámaras Wifi","Cámara Wifi Outdoor Tapo C500 - 2MP - Audio Bidireccional - Visión Nocturna - Detec. Movimiento - MicroSD 512GB - IP65","unidad",401),
    ("TAPO-C510W","Tapo - Cámaras Wifi","Cámara Wifi Outdoor Tapo C510W - 3MP - Audio Bidireccional - Visión Nocturna - Detec. Movimiento - MicroSD 512GB - IP65","unidad",425),
    ("TAPO-L510E","Tapo - Smart","Bombilla Smart Wifi Regulable Tapo L510E - Control Iluminación - Agenda/Temporizador - Control de Voz - Modo Ausente - 8.7W - 806 Lúmenes - Alexa","unidad",87),
    ("TAPO-P105","Tapo - Smart","Mini Enchufe Smart Wifi Tapo P105 1-Pack - Wifi 2.4GHz/BT 4.2 - Temporizador - Modo Ausente - Alexa y Google Assistant","unidad",97),
    ("TAPO-P100","Tapo - Smart","Mini Enchufe Smart Wifi Tapo P100 1-Pack - Wifi 2.4GHz/BT 4.2 - Temporizador - Modo Ausente - Alexa y Google Assistant","unidad",109),
]

@app.get("/admin/cargar-camaras-wifi", response_class=HTMLResponse)
def cargar_camaras_wifi(request: Request, confirmar: str = ""):
    gate = require_login(request)
    if gate:
        return gate

    if confirmar != "si":
        return HTMLResponse("""
        <html><body style="font-family:sans-serif;padding:30px;max-width:500px;margin:auto;">
        <h3>Carga masiva: Cámaras Wifi Enero 2026</h3>
        <p>Se cargarán <strong>68 productos</strong> (IMOU, Dahua, Smart Wifi, Tapo).</p>
        <p>Los productos con SKU duplicado serán <strong>omitidos</strong> automáticamente.</p>
        <a href="/admin/cargar-camaras-wifi?confirmar=si"
           style="background:#198754;color:white;padding:12px 24px;border-radius:6px;text-decoration:none;display:inline-block;margin-top:10px;">
           Confirmar y cargar productos
        </a>
        &nbsp;
        <a href="/productos" style="color:#666;">Cancelar</a>
        </body></html>
        """)

    insertados = 0
    omitidos = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sku, categoria, nombre, unidad, precio in CAMARAS_WIFI:
        if IS_POSTGRES:
            existe = db_fetchone("SELECT id FROM products WHERE sku=%s", (sku,))
            if not existe:
                db_exec(
                    "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo) VALUES(%s,%s,%s,%s,%s,1)",
                    (sku, categoria, nombre, unidad, precio),
                )
                insertados += 1
            else:
                omitidos += 1
        else:
            existe = db_fetchone("SELECT id FROM products WHERE sku=?", (sku,))
            if not existe:
                db_exec(
                    "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo, created_at) VALUES(?,?,?,?,?,1,?)",
                    (sku, categoria, nombre, unidad, precio, now),
                )
                insertados += 1
            else:
                omitidos += 1

    return HTMLResponse(f"""
    <html><body style="font-family:sans-serif;padding:30px;max-width:500px;margin:auto;">
    <h3>Carga completada</h3>
    <p><strong>{insertados}</strong> productos insertados correctamente.</p>
    <p><strong>{omitidos}</strong> productos omitidos (SKU ya existía).</p>
    <a href="/productos" style="background:#0d6efd;color:white;padding:12px 24px;border-radius:6px;text-decoration:none;display:inline-block;margin-top:10px;">
       Ver productos
    </a>
    </body></html>
    """)


PRODUCTOS_VARIOS = [
    ("Z050","Networking - Switches","Switch POE 4 Puertos - 2 Uplink 10/100M - IEEE 802.3/802.3u/802.3af/a - Dist. TX POE 250M con UTP Cat6 - Dimensiones 200x120x45mm","unidad",215.00),
    ("DH-CS4218-18ET-135","Networking - Switches","Switch Administrable 18 Puertos Metálico - 16P POE Potencia 90W - Dist. Max UTP 250M - POE Watchdog - 16P RJ45 10/100Mbps - 2P RJ45 10/100/1000Mbps - 2P SFP","unidad",1744.00),
    ("DHI-ASA1222GL-D","Control de Asistencia","Control Asistencia Standalone - Pantalla 2.4 - 1000 Usuarios - 2000 Huellas - 1000 Passwords - 1000 Tarjetas - Importa/Exporta datos USB Flash","unidad",685.00),
    ("DHI-ASA1222EL-S","Control de Asistencia","Control Asistencia Standalone - Pantalla 2.4 - 1000 Usuarios - 2000 Huellas - 1000 Passwords - Importa/Exporta datos USB Flash - 1 RJ45","unidad",441.00),
    ("SenseFP-M1","Control de Asistencia","Control Asistencia - Pantalla 2.8 - Huella 500 - RFID 500 125kHz - Cap. Reg. 100000 - USB Host - Wifi - Software ZLink - Incluye Fuente","unidad",692.00),
    ("M1-FP-RFID","Control de Asistencia","Control de Asistencia - Pantalla 2.8 - Teclado Físico - RFID 1000 - Huella 1000 - Wifi - Software BioTime Cloud/ZKBio CVAccess - Incluye Fuente de Poder","unidad",585.00),
    ("F302","Accesorios - Baluns","Balun Pasivo 4 en 1 - Transmisor Video HDCVI/AHD/Análogo/TVI - Distancia Max 440M en 720P - Reducción de Ruido","unidad",15.00),
    ("Z043","Accesorios - Cajas","Caja de Conexiones para Cámaras Cuadrada - Material ABS - Temperatura -10 a 60C - IP44 - Dimensiones 11.8x10x5cm","unidad",12.00),
    ("C723","Cableado - UTP","Cable UTP Categoría 6 CCAE 80% Cobre - 4 Pares - 2x0.57mm 23AWG 75 grados - Presentación 305 Metros - Color Blanco - Bajo Norma UL","rollo",604.00),
    ("C712","Cableado - Conectores","Conectores RJ45 Categoría 6 - Diámetro Inserción para Cable 7.8mm - Paquete de 50 Unidades","paquete",16.00),
    ("HUA722020ALA331","Almacenamiento - Discos","Disco Duro New Pull 2TB Ultrastar Hitachi - Cache 32MB - 7200RPM - SATA 3.0Gb/s - Enterprise","unidad",423.00),
    ("SE-R9U-5409A","Infraestructura - Gabinetes","Gabinete 9U 19 Negro - Puerta de Vidrio","unidad",600.00),
    ("IPC-HDW1230T1P-0280B-S6","Cámaras IP - Domo","Cámara IP Domo 2MP Metal/Plástica - Lente 2.8mm - Sensor 1/2.8 CMOS - IR 30M - ROI - Smart H.265+/H.264+ - DWDR - 3D NR - IP67 - 12V/PoE","unidad",385.00),
    ("DHI-NVR2116HS-4KS3","Grabadores - NVR","Grabador IP 16 Canales - Protec. Perimetral - SMD Plus - 1VGA - 1HDMI - 1HDD hasta 16TB - 1RJ45 - ONVIF - Resolución 12MP/8MP/5MP/4MP/3MP/2MP/720p","unidad",745.00),
]

@app.get("/admin/cargar-productos-varios", response_class=HTMLResponse)
def cargar_productos_varios(request: Request, confirmar: str = ""):
    gate = require_login(request)
    if gate:
        return gate

    if confirmar != "si":
        return HTMLResponse("""
        <html><body style="font-family:sans-serif;padding:30px;max-width:500px;margin:auto;">
        <h3>Carga masiva: Productos Varios (Feb/Mar 2026)</h3>
        <p>Se cargarán <strong>14 productos</strong>: Switches POE, Controles de Asistencia, Accesorios, Cableado, Almacenamiento, Grabadores.</p>
        <p>Los productos con SKU duplicado serán <strong>omitidos</strong> automáticamente.</p>
        <a href="/admin/cargar-productos-varios?confirmar=si"
           style="background:#198754;color:white;padding:12px 24px;border-radius:6px;text-decoration:none;display:inline-block;margin-top:10px;">
           Confirmar y cargar productos
        </a>
        &nbsp;
        <a href="/productos" style="color:#666;">Cancelar</a>
        </body></html>
        """)

    insertados = 0
    omitidos = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for sku, categoria, nombre, unidad, precio in PRODUCTOS_VARIOS:
        if IS_POSTGRES:
            existe = db_fetchone("SELECT id FROM products WHERE sku=%s", (sku,))
            if not existe:
                db_exec(
                    "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo) VALUES(%s,%s,%s,%s,%s,1)",
                    (sku, categoria, nombre, unidad, precio),
                )
                insertados += 1
            else:
                omitidos += 1
        else:
            existe = db_fetchone("SELECT id FROM products WHERE sku=?", (sku,))
            if not existe:
                db_exec(
                    "INSERT INTO products(sku, categoria, nombre, unidad, precio_bs, activo, created_at) VALUES(?,?,?,?,?,1,?)",
                    (sku, categoria, nombre, unidad, precio, now),
                )
                insertados += 1
            else:
                omitidos += 1

    return HTMLResponse(f"""
    <html><body style="font-family:sans-serif;padding:30px;max-width:500px;margin:auto;">
    <h3>Carga completada</h3>
    <p><strong>{insertados}</strong> productos insertados correctamente.</p>
    <p><strong>{omitidos}</strong> productos omitidos (SKU ya existía).</p>
    <a href="/productos" style="background:#0d6efd;color:white;padding:12px 24px;border-radius:6px;text-decoration:none;display:inline-block;margin-top:10px;">
       Ver productos
    </a>
    </body></html>
    """)
@dataclass
class InstalacionRow:
    id: int
    quote_id: int
    quote_no: int
    client_name: str
    fecha_instalacion: str
    tecnico: str
    estado: str
    notas_instalacion: str
    total_con_iva: float

def init_instalaciones_table():
    if IS_POSTGRES:
        db_exec("""
            CREATE TABLE IF NOT EXISTS instalaciones (
                id SERIAL PRIMARY KEY,
                quote_id INTEGER NOT NULL REFERENCES quotes(id) ON DELETE CASCADE,
                fecha_instalacion DATE NOT NULL,
                tecnico TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'pendiente',
                notas_instalacion TEXT,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
        # Tabla técnicos
        db_exec("""
            CREATE TABLE IF NOT EXISTS tecnicos (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                telefono TEXT,
                especialidad TEXT,
                activo INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
        # Tabla relación instalacion <-> técnicos (múltiples)
        db_exec("""
            CREATE TABLE IF NOT EXISTS instalacion_tecnicos (
                id SERIAL PRIMARY KEY,
                instalacion_id INTEGER NOT NULL REFERENCES instalaciones(id) ON DELETE CASCADE,
                tecnico_id INTEGER,
                tecnico_nombre TEXT NOT NULL
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS gastos_trabajo (
                id SERIAL PRIMARY KEY,
                quote_id INTEGER NOT NULL REFERENCES quotes(id) ON DELETE CASCADE,
                categoria TEXT NOT NULL,
                descripcion TEXT NOT NULL,
                monto NUMERIC(12,2) NOT NULL,
                created_at TIMESTAMP NOT NULL DEFAULT NOW()
            )
        """)
    else:
        db_exec("""
            CREATE TABLE IF NOT EXISTS instalaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_id INTEGER NOT NULL,
                fecha_instalacion TEXT NOT NULL,
                tecnico TEXT NOT NULL,
                estado TEXT NOT NULL DEFAULT 'pendiente',
                notas_instalacion TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (quote_id) REFERENCES quotes(id)
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS tecnicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                telefono TEXT,
                especialidad TEXT,
                activo INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS instalacion_tecnicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instalacion_id INTEGER NOT NULL,
                tecnico_id INTEGER,
                tecnico_nombre TEXT NOT NULL,
                FOREIGN KEY (instalacion_id) REFERENCES instalaciones(id)
            )
        """)
        db_exec("""
            CREATE TABLE IF NOT EXISTS gastos_trabajo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_id INTEGER NOT NULL,
                categoria TEXT NOT NULL,
                descripcion TEXT NOT NULL,
                monto REAL NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (quote_id) REFERENCES quotes(id)
            )
        """)

init_instalaciones_table()

def get_quote_total(quote_id: int) -> float:
    if IS_POSTGRES:
        rows = db_fetchall("SELECT qty, unit_price FROM quote_items WHERE quote_id=%s", (quote_id,))
    else:
        rows = db_fetchall("SELECT qty, unit_price FROM quote_items WHERE quote_id=?", (quote_id,))
    subtotal = sum(float(r["qty"]) * float(r["unit_price"]) for r in rows)
    return subtotal * (1 + IVA_RATE)

# =========================
# Técnicos helpers y rutas
# =========================
@dataclass
class TecnicoRow:
    id: int
    nombre: str
    telefono: str
    especialidad: str
    activo: int
    total_instalaciones: int = 0

def load_tecnicos_activos() -> List[TecnicoRow]:
    rows = db_fetchall("SELECT id, nombre, telefono, especialidad, activo FROM tecnicos WHERE activo=1 ORDER BY nombre")
    return [TecnicoRow(id=int(r["id"]), nombre=str(r["nombre"]), telefono=str(r["telefono"] or ""),
                       especialidad=str(r["especialidad"] or ""), activo=1) for r in rows]

def load_all_tecnicos() -> List[TecnicoRow]:
    if IS_POSTGRES:
        rows = db_fetchall("""
            SELECT t.id, t.nombre, t.telefono, t.especialidad, t.activo,
                   COUNT(it.id) AS total_instalaciones
            FROM tecnicos t
            LEFT JOIN instalacion_tecnicos it ON it.tecnico_id = t.id
            GROUP BY t.id, t.nombre, t.telefono, t.especialidad, t.activo
            ORDER BY t.nombre
        """)
    else:
        rows = db_fetchall("""
            SELECT t.id, t.nombre, t.telefono, t.especialidad, t.activo,
                   COUNT(it.id) AS total_instalaciones
            FROM tecnicos t
            LEFT JOIN instalacion_tecnicos it ON it.tecnico_id = t.id
            GROUP BY t.id
            ORDER BY t.nombre
        """)
    return [TecnicoRow(id=int(r["id"]), nombre=str(r["nombre"]), telefono=str(r["telefono"] or ""),
                       especialidad=str(r["especialidad"] or ""), activo=int(r["activo"]),
                       total_instalaciones=int(r["total_instalaciones"])) for r in rows]

@app.get("/tecnicos", response_class=HTMLResponse)
def tecnicos_get(request: Request):
    gate = require_login(request)
    if gate:
        return gate
    return templates.TemplateResponse("tecnicos.html", {
        "request": request,
        "empresa": EMPRESA_NOMBRE,
        "tecnicos": load_all_tecnicos(),
        "msg": request.query_params.get("msg", ""),
        "msg_type": request.query_params.get("msg_type", "success"),
    })

@app.post("/tecnicos/guardar")
def tecnicos_guardar(
    request: Request,
    tecnico_id: str = Form(""),
    nombre: str = Form(...),
    telefono: str = Form(""),
    especialidad: str = Form(""),
    activo: int = Form(1),
):
    gate = require_login(request)
    if gate:
        return gate

    nombre = nombre.strip()
    telefono = telefono.strip() or None
    especialidad = especialidad.strip() or None

    if tecnico_id:
        if IS_POSTGRES:
            db_exec("UPDATE tecnicos SET nombre=%s, telefono=%s, especialidad=%s, activo=%s WHERE id=%s",
                    (nombre, telefono, especialidad, activo, int(tecnico_id)))
        else:
            db_exec("UPDATE tecnicos SET nombre=?, telefono=?, especialidad=?, activo=? WHERE id=?",
                    (nombre, telefono, especialidad, activo, int(tecnico_id)))
        msg = "Técnico+actualizado."
    else:
        if IS_POSTGRES:
            db_exec("INSERT INTO tecnicos(nombre, telefono, especialidad, activo) VALUES(%s,%s,%s,%s)",
                    (nombre, telefono, especialidad, activo))
        else:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db_exec("INSERT INTO tecnicos(nombre, telefono, especialidad, activo, created_at) VALUES(?,?,?,?,?)",
                    (nombre, telefono, especialidad, activo, now))
        msg = "Técnico+agregado."

    return RedirectResponse(url=f"/tecnicos?msg={msg}&msg_type=success", status_code=303)

@app.post("/tecnicos/{tecnico_id}/borrar")
def tecnicos_borrar(request: Request, tecnico_id: int):
    gate = require_login(request)
    if gate:
        return gate
    if IS_POSTGRES:
        db_exec("DELETE FROM tecnicos WHERE id=%s", (tecnico_id,))
    else:
        db_exec("DELETE FROM tecnicos WHERE id=?", (tecnico_id,))
    return RedirectResponse(url="/tecnicos?msg=Técnico+eliminado.&msg_type=warning", status_code=303)


@app.get("/instalacion/{quote_id}/agendar", response_class=HTMLResponse)
def agendar_get(request: Request, quote_id: int):
    gate = require_login(request)
    if gate:
        return gate

    if IS_POSTGRES:
        q = db_fetchone("SELECT * FROM quotes WHERE id=%s", (quote_id,))
        inst = db_fetchone("SELECT * FROM instalaciones WHERE quote_id=%s", (quote_id,))
    else:
        q = db_fetchone("SELECT * FROM quotes WHERE id=?", (quote_id,))
        inst = db_fetchone("SELECT * FROM instalaciones WHERE quote_id=?", (quote_id,))

    if not q:
        return RedirectResponse(url="/historial", status_code=303)

    # Técnicos ya asignados a esta instalación
    tecnicos_asignados = []
    if inst:
        if IS_POSTGRES:
            tecs = db_fetchall("SELECT tecnico_id, tecnico_nombre FROM instalacion_tecnicos WHERE instalacion_id=%s", (inst["id"],))
        else:
            tecs = db_fetchall("SELECT tecnico_id, tecnico_nombre FROM instalacion_tecnicos WHERE instalacion_id=?", (inst["id"],))
        tecnicos_asignados = [{"id": r["tecnico_id"], "nombre": r["tecnico_nombre"]} for r in tecs]

    # Técnicos activos para el select
    tecs_activos = load_tecnicos_activos()

    import json
    return templates.TemplateResponse("agendar.html", {
        "request": request,
        "q": q,
        "instalacion": inst,
        "empresa": EMPRESA_NOMBRE,
        "tecnicos_activos": tecs_activos,
        "tecnicos_asignados_json": json.dumps(tecnicos_asignados),
    })

@app.post("/instalacion/{quote_id}/guardar")
def agendar_post(
    request: Request,
    quote_id: int,
    fecha_instalacion: str = Form(...),
    tecnicos_json: str = Form("[]"),
    estado: str = Form("pendiente"),
    notas_instalacion: str = Form(""),
):
    gate = require_login(request)
    if gate:
        return gate

    import json as _json
    try:
        tecs = _json.loads(tecnicos_json)
    except Exception:
        tecs = []

    # Nombre resumido para campo legado "tecnico"
    tecnico_str = ", ".join(t["nombre"] for t in tecs) if tecs else "Sin asignar"

    if IS_POSTGRES:
        existing = db_fetchone("SELECT id FROM instalaciones WHERE quote_id=%s", (quote_id,))
        if existing:
            inst_id = int(existing["id"])
            db_exec(
                "UPDATE instalaciones SET fecha_instalacion=%s, tecnico=%s, estado=%s, notas_instalacion=%s WHERE id=%s",
                (fecha_instalacion, tecnico_str, estado, notas_instalacion.strip() or None, inst_id),
            )
            db_exec("DELETE FROM instalacion_tecnicos WHERE instalacion_id=%s", (inst_id,))
        else:
            con = db_connect()
            try:
                cur = con.cursor()
                cur.execute(
                    "INSERT INTO instalaciones(quote_id, fecha_instalacion, tecnico, estado, notas_instalacion) VALUES(%s,%s,%s,%s,%s) RETURNING id",
                    (quote_id, fecha_instalacion, tecnico_str, estado, notas_instalacion.strip() or None),
                )
                inst_id = int(cur.fetchone()["id"])
                con.commit()
            finally:
                con.close()
        for t in tecs:
            db_exec(
                "INSERT INTO instalacion_tecnicos(instalacion_id, tecnico_id, tecnico_nombre) VALUES(%s,%s,%s)",
                (inst_id, t.get("id"), t["nombre"]),
            )
    else:
        existing = db_fetchone("SELECT id FROM instalaciones WHERE quote_id=?", (quote_id,))
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if existing:
            inst_id = int(existing["id"])
            db_exec(
                "UPDATE instalaciones SET fecha_instalacion=?, tecnico=?, estado=?, notas_instalacion=? WHERE id=?",
                (fecha_instalacion, tecnico_str, estado, notas_instalacion.strip() or None, inst_id),
            )
            db_exec("DELETE FROM instalacion_tecnicos WHERE instalacion_id=?", (inst_id,))
        else:
            con = db_connect()
            try:
                cur = con.cursor()
                cur.execute(
                    "INSERT INTO instalaciones(quote_id, fecha_instalacion, tecnico, estado, notas_instalacion, created_at) VALUES(?,?,?,?,?,?)",
                    (quote_id, fecha_instalacion, tecnico_str, estado, notas_instalacion.strip() or None, now),
                )
                inst_id = cur.lastrowid
                con.commit()
            finally:
                con.close()
        for t in tecs:
            db_exec(
                "INSERT INTO instalacion_tecnicos(instalacion_id, tecnico_id, tecnico_nombre) VALUES(?,?,?)",
                (inst_id, t.get("id"), t["nombre"]),
            )

    return RedirectResponse(url="/agenda", status_code=303)

@app.get("/agenda", response_class=HTMLResponse)
def agenda(request: Request, fecha: str = ""):
    gate = require_login(request)
    if gate:
        return gate

    fecha_sel = fecha or datetime.now().strftime("%Y-%m-%d")

    if IS_POSTGRES:
        rows_dia = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado, i.notas_instalacion,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion = %s
            ORDER BY i.tecnico
        """, (fecha_sel,))
        rows_proximas = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion > %s AND i.estado != 'completada'
            ORDER BY i.fecha_instalacion
            LIMIT 20
        """, (fecha_sel,))
    else:
        rows_dia = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado, i.notas_instalacion,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion = ?
            ORDER BY i.tecnico
        """, (fecha_sel,))
        rows_proximas = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion > ? AND i.estado != 'completada'
            ORDER BY i.fecha_instalacion
            LIMIT 20
        """, (fecha_sel,))

    def enrich(rows):
        result = []
        for r in rows:
            total = get_quote_total(int(r["quote_id"]))
            result.append({
                "quote_id": r["quote_id"],
                "quote_no": r["quote_no"],
                "client_name": r["client_name"],
                "fecha_instalacion": str(r["fecha_instalacion"]),
                "tecnico": r["tecnico"],
                "estado": r["estado"],
                "notas_instalacion": r.get("notas_instalacion", ""),
                "total_con_iva": total,
            })
        return result

    instalaciones_dia = enrich(rows_dia)
    proximas = enrich(rows_proximas)

    return templates.TemplateResponse("agenda.html", {
        "request": request,
        "empresa": EMPRESA_NOMBRE,
        "fecha_sel": fecha_sel,
        "instalaciones_dia": instalaciones_dia,
        "proximas": proximas,
        "total_dia": len(instalaciones_dia),
    })


# =========================
# Gastos por trabajo
# =========================
@dataclass
class GastoRow:
    id: int
    quote_id: int
    categoria: str
    descripcion: str
    monto: float

def get_gastos(quote_id: int) -> List[GastoRow]:
    if IS_POSTGRES:
        rows = db_fetchall("SELECT id, quote_id, categoria, descripcion, monto FROM gastos_trabajo WHERE quote_id=%s ORDER BY id", (quote_id,))
    else:
        rows = db_fetchall("SELECT id, quote_id, categoria, descripcion, monto FROM gastos_trabajo WHERE quote_id=? ORDER BY id", (quote_id,))
    return [GastoRow(id=int(r["id"]), quote_id=int(r["quote_id"]), categoria=str(r["categoria"]),
                     descripcion=str(r["descripcion"]), monto=float(r["monto"])) for r in rows]

def get_total_gastos(quote_id: int) -> float:
    gastos = get_gastos(quote_id)
    return sum(g.monto for g in gastos)

@app.get("/gastos/{quote_id}", response_class=HTMLResponse)
def gastos_get(request: Request, quote_id: int, msg: str = "", msg_type: str = "success"):
    gate = require_login(request)
    if gate:
        return gate
    if IS_POSTGRES:
        q = db_fetchone("SELECT id, quote_no, client_name, created_at FROM quotes WHERE id=%s", (quote_id,))
    else:
        q = db_fetchone("SELECT id, quote_no, client_name, created_at FROM quotes WHERE id=?", (quote_id,))
    if not q:
        return RedirectResponse(url="/historial", status_code=303)

    gastos = get_gastos(quote_id)
    total_cotizacion = get_quote_total(quote_id)
    total_gastos = sum(g.monto for g in gastos)
    utilidad = total_cotizacion - total_gastos
    margen = (utilidad / total_cotizacion * 100) if total_cotizacion > 0 else 0

    desglose: dict = {}
    for g in gastos:
        desglose[g.categoria] = desglose.get(g.categoria, 0) + g.monto

    return templates.TemplateResponse("gastos.html", {
        "request": request,
        "empresa": EMPRESA_NOMBRE,
        "q": dict(q),
        "gastos": gastos,
        "total_cotizacion": total_cotizacion,
        "total_gastos": total_gastos,
        "utilidad": utilidad,
        "margen": margen,
        "desglose": desglose,
        "msg": msg,
        "msg_type": msg_type,
    })

@app.post("/gastos/{quote_id}/agregar")
def gastos_agregar(request: Request, quote_id: int,
                   categoria: str = Form(...), descripcion: str = Form(...), monto: float = Form(...)):
    gate = require_login(request)
    if gate:
        return gate
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if IS_POSTGRES:
        db_exec("INSERT INTO gastos_trabajo(quote_id, categoria, descripcion, monto) VALUES(%s,%s,%s,%s)",
                (quote_id, categoria, descripcion, monto))
    else:
        db_exec("INSERT INTO gastos_trabajo(quote_id, categoria, descripcion, monto, created_at) VALUES(?,?,?,?,?)",
                (quote_id, categoria, descripcion, monto, now))
    return RedirectResponse(url=f"/gastos/{quote_id}?msg=Gasto+agregado.&msg_type=success", status_code=303)

@app.post("/gastos/{gasto_id}/borrar")
def gastos_borrar(request: Request, gasto_id: int):
    gate = require_login(request)
    if gate:
        return gate
    if IS_POSTGRES:
        g = db_fetchone("SELECT quote_id FROM gastos_trabajo WHERE id=%s", (gasto_id,))
        db_exec("DELETE FROM gastos_trabajo WHERE id=%s", (gasto_id,))
    else:
        g = db_fetchone("SELECT quote_id FROM gastos_trabajo WHERE id=?", (gasto_id,))
        db_exec("DELETE FROM gastos_trabajo WHERE id=?", (gasto_id,))
    quote_id = int(g["quote_id"]) if g else 0
    return RedirectResponse(url=f"/gastos/{quote_id}?msg=Gasto+eliminado.&msg_type=warning", status_code=303)

@app.get("/reportes", response_class=HTMLResponse)
def reportes(request: Request, desde: str = "", hasta: str = "", tecnico: str = ""):
    gate = require_login(request)
    if gate:
        return gate

    hoy = datetime.now().strftime("%Y-%m-%d")
    desde = desde or datetime.now().strftime("%Y-%m-01")
    hasta = hasta or hoy

    if IS_POSTGRES:
        rows = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion BETWEEN %s AND %s
            ORDER BY i.fecha_instalacion
        """, (desde, hasta))
        tecnicos_rows = db_fetchall("SELECT DISTINCT tecnico FROM instalaciones ORDER BY tecnico")
    else:
        rows = db_fetchall("""
            SELECT i.quote_id, i.fecha_instalacion, i.tecnico, i.estado,
                   q.quote_no, q.client_name
            FROM instalaciones i
            JOIN quotes q ON q.id = i.quote_id
            WHERE i.fecha_instalacion BETWEEN ? AND ?
            ORDER BY i.fecha_instalacion
        """, (desde, hasta))
        tecnicos_rows = db_fetchall("SELECT DISTINCT tecnico FROM instalaciones ORDER BY tecnico")

    tecnicos = [r["tecnico"] for r in tecnicos_rows]

    instalaciones = []
    for r in rows:
        if tecnico and r["tecnico"] != tecnico:
            continue
        total = get_quote_total(int(r["quote_id"]))
        total_gastos = get_total_gastos(int(r["quote_id"]))
        utilidad = total - total_gastos
        margen = (utilidad / total * 100) if total > 0 else 0
        instalaciones.append({
            "quote_id": r["quote_id"],
            "quote_no": r["quote_no"],
            "client_name": r["client_name"],
            "fecha_instalacion": str(r["fecha_instalacion"]),
            "tecnico": r["tecnico"],
            "estado": r["estado"],
            "total_con_iva": total,
            "total_gastos": total_gastos,
            "utilidad": utilidad,
            "margen": margen,
        })

    total_con_iva = sum(i["total_con_iva"] for i in instalaciones)
    total_sin_iva = total_con_iva / (1 + IVA_RATE)
    total_iva = total_con_iva - total_sin_iva
    total_gastos_global = sum(i["total_gastos"] for i in instalaciones)
    utilidad_global = total_con_iva - total_gastos_global
    margen_global = (utilidad_global / total_con_iva * 100) if total_con_iva > 0 else 0

    stats = {
        "total": len(instalaciones),
        "pendientes": sum(1 for i in instalaciones if i["estado"] == "pendiente"),
        "en_curso": sum(1 for i in instalaciones if i["estado"] == "en_curso"),
        "completadas": sum(1 for i in instalaciones if i["estado"] == "completada"),
        "total_sin_iva": total_sin_iva,
        "total_iva": total_iva,
        "total_con_iva": total_con_iva,
        "total_gastos": total_gastos_global,
        "utilidad": utilidad_global,
        "margen": margen_global,
    }

    return templates.TemplateResponse("reportes.html", {
        "request": request,
        "empresa": EMPRESA_NOMBRE,
        "desde": desde,
        "hasta": hasta,
        "tecnico_filtro": tecnico,
        "tecnicos": tecnicos,
        "instalaciones": instalaciones,
        "stats": stats,
    })