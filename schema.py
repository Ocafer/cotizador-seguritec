from __future__ import annotations
import os
from datetime import datetime
from typing import List

from openpyxl import load_workbook

from config import EXCEL_PATH
from database import db_connect, db_exec, db_fetchone, IS_POSTGRES, psql
from models import Product


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def _norm_header(x) -> str:
    return str(x).strip().lower() if x is not None else ""


def _to_float(x, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        return float(x)
    except Exception:
        s = str(x).strip().replace(",", ".")
        try:
            return float(s)
        except Exception:
            return default


def _to_int(x, default: int = 0) -> int:
    try:
        if x is None:
            return default
        return int(float(x))
    except Exception:
        return default


def read_products_from_excel() -> List[Product]:
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "productos" not in wb.sheetnames:
        raise ValueError("El Excel debe tener una hoja llamada 'productos'.")
    ws = wb["productos"]
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        h = _norm_header(cell.value)
        if h:
            headers[h] = col_idx
    required = {"sku", "categoria", "nombre", "unidad", "precio_bs", "activo"}
    missing = required - set(headers.keys())
    if missing:
        raise ValueError(f"Faltan columnas en Excel: {', '.join(sorted(missing))}")
    products: List[Product] = []
    for r in range(2, ws.max_row + 1):
        activo = _to_int(ws.cell(r, headers["activo"]).value, default=0)
        if activo != 1:
            continue
        sku = str(ws.cell(r, headers["sku"]).value or "").strip()
        categoria = str(ws.cell(r, headers["categoria"]).value or "").strip()
        nombre = str(ws.cell(r, headers["nombre"]).value or "").strip()
        unidad = str(ws.cell(r, headers["unidad"]).value or "").strip() or "unidad"
        precio_bs = _to_float(ws.cell(r, headers["precio_bs"]).value, default=0.0)
        if not nombre and not sku:
            continue
        products.append(Product(sku=sku, categoria=categoria, nombre=nombre,
                                unidad=unidad, precio_bs=precio_bs, activo=1))
    products.sort(key=lambda p: (p.categoria.lower(), p.nombre.lower()))
    return products


# ---------------------------------------------------------------------------
# Seeding & sequence
# ---------------------------------------------------------------------------

def products_count() -> int:
    row = db_fetchone("SELECT COUNT(*) AS total FROM products")
    return int(row["total"]) if row else 0


def seed_products_from_excel_if_empty() -> None:
    if products_count() > 0:
        return
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        products = read_products_from_excel()
    except Exception:
        return
    if not products:
        return
    if IS_POSTGRES:
        for p in products:
            db_exec(
                psql("""INSERT INTO products(sku,categoria,nombre,unidad,precio_bs,activo)
                        VALUES(?,?,?,?,?,?) ON CONFLICT (sku) DO UPDATE SET
                        categoria=EXCLUDED.categoria, nombre=EXCLUDED.nombre,
                        unidad=EXCLUDED.unidad, precio_bs=EXCLUDED.precio_bs, activo=EXCLUDED.activo"""),
                (p.sku or None, p.categoria, p.nombre, p.unidad, p.precio_bs, p.activo),
            )
    else:
        con = db_connect()
        try:
            cur = con.cursor()
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for p in products:
                cur.execute(
                    "INSERT OR REPLACE INTO products(sku,categoria,nombre,unidad,precio_bs,activo,created_at) VALUES(?,?,?,?,?,?,?)",
                    (p.sku or None, p.categoria, p.nombre, p.unidad, p.precio_bs, p.activo, now),
                )
            con.commit()
        finally:
            con.close()


def next_quote_no() -> int:
    if IS_POSTGRES:
        db_exec("UPDATE counter SET last_quote_no = last_quote_no + 1")
        row = db_fetchone("SELECT last_quote_no FROM counter LIMIT 1")
        return int(row["last_quote_no"])
    else:
        con = db_connect()
        try:
            cur = con.cursor()
            cur.execute("UPDATE counter SET value = value + 1 WHERE key='quote_no'")
            cur.execute("SELECT value FROM counter WHERE key='quote_no'")
            n = cur.fetchone()[0]
            con.commit()
            return int(n)
        finally:
            con.close()
